from __future__ import annotations

from dataclasses import dataclass
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.document.adapters.base import BaseDocumentAdapter


@dataclass
class ExcelCellData:
    row: int
    col: int
    coordinate: str
    value: Any


class ExcelAdapter(BaseDocumentAdapter):
    def load(self, file_path: str) -> Workbook:
        return load_workbook(file_path)

    def save(self, document: Workbook, output_path: str) -> str:
        document.save(output_path)
        return output_path

    def list_sheets(self, workbook: Workbook) -> list[str]:
        return workbook.sheetnames

    def get_sheet(self, workbook: Workbook, sheet_name: str | None = None) -> Worksheet:
        if sheet_name:
            return workbook[sheet_name]
        return workbook[workbook.sheetnames[0]]

    def read_used_range(self, worksheet: Worksheet) -> list[list[Any]]:
        data: list[list[Any]] = []
        for row in worksheet.iter_rows(values_only=True):
            data.append(list(row))
        return data

    def read_range(self, worksheet: Worksheet, cell_range: str) -> list[list[Any]]:
        result: list[list[Any]] = []
        for row in worksheet[cell_range]:
            result.append([cell.value for cell in row])
        return result

    def write_cell(self, worksheet: Worksheet, cell_ref: str, value: Any) -> None:
        worksheet[cell_ref] = value

    def write_rows(self, worksheet: Worksheet, start_row: int, start_col: int, rows: list[list[Any]]) -> None:
        for r_offset, row in enumerate(rows):
            for c_offset, value in enumerate(row):
                worksheet.cell(row=start_row + r_offset, column=start_col + c_offset, value=value)

    def find_header_row(self, worksheet: Worksheet, expected_headers: list[str], scan_rows: int = 10) -> int | None:
        normalized_expected = [str(h).strip() for h in expected_headers]
        for row_idx in range(1, scan_rows + 1):
            values = []
            for cell in worksheet[row_idx]:
                values.append("" if cell.value is None else str(cell.value).strip())
            if all(h in values for h in normalized_expected):
                return row_idx
        return None

    def find_last_data_row(
        self,
        worksheet: Worksheet,
        header_row: int,
        key_col: int = 1,
    ) -> int:
        """从 header_row 向下扫描，返回最后一个非空数据行的行号。"""
        last_row = header_row
        for cell in worksheet.iter_rows(
            min_row=header_row + 1,
            min_col=key_col,
            max_col=key_col,
            values_only=False,
        ):
            if cell[0].value not in (None, ""):
                last_row = cell[0].row
        return last_row

    def append_rows_by_header(
        self,
        worksheet: Worksheet,
        expected_headers: list[str],
        rows: list[dict],
    ) -> int:
        """
        找到表头行，在表格末尾追加数据行。
        rows: list of dicts keyed by header name.
        返回写入行数。
        """
        header_row = self.find_header_row(worksheet, expected_headers)
        if header_row is None:
            return 0

        # 建立表头→列索引映射
        header_map: dict[str, int] = {}
        for cell in worksheet[header_row]:
            value = "" if cell.value is None else str(cell.value).strip()
            if value in expected_headers:
                header_map[value] = cell.column

        last_row = self.find_last_data_row(worksheet, header_row)
        start_row = last_row + 1

        for r_offset, record in enumerate(rows):
            for header, col_idx in header_map.items():
                worksheet.cell(
                    row=start_row + r_offset,
                    column=col_idx,
                    value=record.get(header),
                )

        return len(rows)

    def extract_table_by_header(
        self,
        worksheet: Worksheet,
        expected_headers: list[str],
        max_empty_rows: int = 2,
    ) -> list[dict[str, Any]]:
        header_row = self.find_header_row(worksheet, expected_headers)
        if header_row is None:
            return []

        header_map: dict[str, int] = {}
        row_cells = list(worksheet[header_row])
        for idx, cell in enumerate(row_cells, start=1):
            value = "" if cell.value is None else str(cell.value).strip()
            if value in expected_headers:
                header_map[value] = idx

        records: list[dict[str, Any]] = []
        empty_streak = 0
        row_idx = header_row + 1

        while True:
            row_record: dict[str, Any] = {}
            row_has_any_value = False

            for header in expected_headers:
                col_idx = header_map.get(header)
                value = worksheet.cell(row=row_idx, column=col_idx).value if col_idx else None
                row_record[header] = value
                if value not in (None, ""):
                    row_has_any_value = True

            if not row_has_any_value:
                empty_streak += 1
                if empty_streak >= max_empty_rows:
                    break
            else:
                empty_streak = 0
                records.append(row_record)

            row_idx += 1

        return records
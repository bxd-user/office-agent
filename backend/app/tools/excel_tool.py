import csv
import os
from datetime import date, datetime

from openpyxl import load_workbook
from app.tools.base_tool import BaseDataTool


class ExcelTool(BaseDataTool):
    TOOL_NAME = "excel"
    SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm", ".csv", ".tsv"}

    def read_for_llm(self, path: str, **kwargs) -> dict:
        max_sheets = int(kwargs.get("max_sheets", 20))
        max_rows_per_sheet = int(kwargs.get("max_rows_per_sheet", 2000))
        max_cols_per_sheet = int(kwargs.get("max_cols_per_sheet", 200))
        return self.read_workbook_for_llm(
            path=path,
            max_sheets=max_sheets,
            max_rows_per_sheet=max_rows_per_sheet,
            max_cols_per_sheet=max_cols_per_sheet,
        )

    def write_from_llm(self, source_path: str, data: dict, output_path: str, **kwargs) -> str:
        return self.write_workbook_from_llm(source_path=source_path, workbook_data=data, output_path=output_path)

    def execute_llm_instruction(self, instruction: dict) -> dict:
        action = str(instruction.get("action") or "").strip().lower()

        if action == "read":
            path = str(instruction.get("path") or "")
            options = instruction.get("options", {}) if isinstance(instruction.get("options", {}), dict) else {}
            return {
                "action": "read",
                "data": self.read_for_llm(path, **options),
            }

        if action == "write":
            source_path = str(instruction.get("source_path") or "")
            output_path = str(instruction.get("output_path") or "")
            data = instruction.get("data", {}) if isinstance(instruction.get("data", {}), dict) else {}
            return {
                "action": "write",
                "output_path": self.write_from_llm(source_path=source_path, data=data, output_path=output_path),
            }

        if action == "extract_fields":
            path = str(instruction.get("path") or "")
            max_fields = int(instruction.get("max_fields", 300))
            return {
                "action": "extract_fields",
                "fields": self.get_fields_for_llm(path, max_fields=max_fields),
            }

        if action == "read_first_row":
            path = str(instruction.get("path") or "")
            return {
                "action": "read_first_row",
                "row": self.read_first_row_as_dict(path),
            }

        raise ValueError(f"ExcelTool 不支持的 action: {action}")

    def read_workbook_for_llm(
        self,
        path: str,
        max_sheets: int = 20,
        max_rows_per_sheet: int = 2000,
        max_cols_per_sheet: int = 200,
    ) -> dict:
        ext = self._ensure_supported_extension(path)

        if ext in {".csv", ".tsv"}:
            delimiter = "\t" if ext == ".tsv" else ","
            rows = self._read_delimited_all_rows(path, delimiter=delimiter, max_rows=max_rows_per_sheet)
            rows = [self._truncate_row(row, max_cols_per_sheet) for row in rows]
            return {
                "file_type": ext,
                "sheets": [
                    {
                        "name": "Sheet1",
                        "rows": rows,
                    }
                ],
            }

        sheets = []
        wb = load_workbook(path, data_only=True, read_only=True)
        try:
            for ws in wb.worksheets[:max_sheets]:
                rows = []
                for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    if row_index > max_rows_per_sheet:
                        break

                    normalized_row = [self._to_llm_cell(value) for value in row[:max_cols_per_sheet]]
                    if any(cell != "" for cell in normalized_row):
                        rows.append(normalized_row)

                sheets.append(
                    {
                        "name": ws.title,
                        "rows": rows,
                    }
                )
        finally:
            wb.close()

        return {
            "file_type": ext,
            "sheets": sheets,
        }

    def write_workbook_from_llm(self, source_path: str, workbook_data: dict, output_path: str) -> str:
        source_ext = self._ensure_supported_extension(source_path)
        output_ext = self._ensure_supported_extension(output_path)

        if source_ext != output_ext:
            raise ValueError("source_path 与 output_path 的文件类型必须一致")

        sheets = workbook_data.get("sheets", []) if isinstance(workbook_data, dict) else []
        if not isinstance(sheets, list) or not sheets:
            raise ValueError("workbook_data.sheets 必须是非空列表")

        if output_ext in {".csv", ".tsv"}:
            first = sheets[0]
            rows = self._normalize_rows_for_write(first.get("rows", []))
            self._write_delimited_rows(output_path, rows, delimiter="\t" if output_ext == ".tsv" else ",")
            return output_path

        wb = load_workbook(source_path)
        try:
            for sheet_data in sheets:
                sheet_name = str(sheet_data.get("name") or "").strip() or "Sheet1"
                rows = self._normalize_rows_for_write(sheet_data.get("rows", []))

                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.create_sheet(title=sheet_name)

                self._replace_sheet_values(ws, rows)

            wb.save(output_path)
        finally:
            wb.close()

        return output_path

    def read_first_row_as_dict(self, path: str) -> dict:
        ext = self._ensure_supported_extension(path)

        if ext in {".csv", ".tsv"}:
            return self._read_first_row_from_delimited(path, delimiter="\t" if ext == ".tsv" else ",")

        return self._read_first_row_from_workbook(path)

    def get_fields_for_llm(self, path: str, max_fields: int = 300) -> list[str]:
        ext = self._ensure_supported_extension(path)

        if ext in {".csv", ".tsv"}:
            header, _ = self._read_header_and_first_data_from_delimited(
                path,
                delimiter="\t" if ext == ".tsv" else ",",
            )
            return header[:max_fields]

        fields = []
        seen = set()

        wb = load_workbook(path, data_only=True, read_only=True)
        try:
            for ws in wb.worksheets:
                headers, _ = self._extract_header_and_first_data_from_rows(ws.iter_rows(values_only=True))
                for field in headers:
                    normalized = self._normalize_field(field)
                    if not normalized or normalized in seen:
                        continue
                    seen.add(normalized)
                    fields.append(field)
                    if len(fields) >= max_fields:
                        return fields
        finally:
            wb.close()

        return fields

    def _read_first_row_from_workbook(self, path: str) -> dict:
        wb = load_workbook(path, data_only=True, read_only=True)
        try:
            best = {}
            for ws in wb.worksheets:
                headers, values = self._extract_header_and_first_data_from_rows(ws.iter_rows(values_only=True))
                row_dict = self._build_row_dict(headers, values)
                if len(row_dict) > len(best):
                    best = row_dict
            return best
        finally:
            wb.close()

    def _read_first_row_from_delimited(self, path: str, delimiter: str) -> dict:
        headers, values = self._read_header_and_first_data_from_delimited(path, delimiter)
        return self._build_row_dict(headers, values)

    def _read_header_and_first_data_from_delimited(self, path: str, delimiter: str) -> tuple[list[str], list]:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = list(csv.reader(f, delimiter=delimiter))

        headers, values = self._extract_header_and_first_data_from_rows(reader)
        return headers, values

    def _extract_header_and_first_data_from_rows(self, rows) -> tuple[list[str], list]:
        materialized_rows = []
        for row in rows:
            values = list(row)
            if not values:
                continue
            materialized_rows.append(values)
            if len(materialized_rows) >= 200:
                break

        if not materialized_rows:
            return [], []

        header_index = self._detect_header_row_index(materialized_rows)
        headers = materialized_rows[header_index]

        values = []
        for row in materialized_rows[header_index + 1 :]:
            if any(self._normalize_cell(v) != "" for v in row):
                values = row
                break

        return [self._normalize_cell(h) for h in headers], values

    def _detect_header_row_index(self, rows: list[list]) -> int:
        best_index = 0
        best_score = -1

        for index, row in enumerate(rows[:30]):
            normalized = [self._normalize_cell(v) for v in row]
            non_empty = [v for v in normalized if v]
            unique_count = len(set(non_empty))
            score = len(non_empty) + unique_count
            if score > best_score:
                best_score = score
                best_index = index

        return best_index

    def _build_row_dict(self, headers: list[str], values: list) -> dict:
        result = {}
        for index, header in enumerate(headers):
            key = self._normalize_cell(header)
            if not key:
                continue
            value = values[index] if index < len(values) else ""
            result[key] = self._normalize_cell(value)
        return result

    def _normalize_cell(self, value) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _normalize_field(self, value: str) -> str:
        return "".join(str(value or "").split()).lower()

    def _ensure_supported_extension(self, path: str) -> str:
        ext = os.path.splitext(path)[1].lower()
        if ext not in self.SUPPORTED_EXTENSIONS:
            raise ValueError(f"不支持的 Excel 文件类型: {ext or '无扩展名'}")
        return ext

    def _read_delimited_all_rows(self, path: str, delimiter: str, max_rows: int) -> list[list[str]]:
        rows = []
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f, delimiter=delimiter)
            for index, row in enumerate(reader, start=1):
                if index > max_rows:
                    break
                rows.append([self._to_llm_cell(cell) for cell in row])
        return rows

    def _truncate_row(self, row: list[str], max_cols: int) -> list[str]:
        return row[:max_cols]

    def _to_llm_cell(self, value) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.isoformat(sep=" ", timespec="seconds")
        if isinstance(value, date):
            return value.isoformat()
        return str(value)

    def _normalize_rows_for_write(self, rows) -> list[list[str]]:
        if not isinstance(rows, list):
            raise ValueError("rows 必须是二维数组")

        normalized_rows = []
        for row in rows:
            if not isinstance(row, list):
                raise ValueError("rows 中每一行都必须是数组")
            normalized_rows.append(["" if cell is None else str(cell) for cell in row])
        return normalized_rows

    def _replace_sheet_values(self, ws, rows: list[list[str]]) -> None:
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).value = None

        for r_index, row in enumerate(rows, start=1):
            for c_index, cell_value in enumerate(row, start=1):
                ws.cell(row=r_index, column=c_index).value = cell_value

    def _write_delimited_rows(self, path: str, rows: list[list[str]], delimiter: str) -> None:
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f, delimiter=delimiter)
            for row in rows:
                writer.writerow(row)
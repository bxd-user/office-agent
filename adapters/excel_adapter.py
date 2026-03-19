from __future__ import annotations

from typing import Any, Dict, List, Optional, cast

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from tools.base import ToolValidationError
from tools.common_tools import ensure_excel_file as ensure_excel_file_common


SUPPORTED_EXCEL_EXTENSIONS = (".xlsx", ".xlsm")


def ensure_excel_file(file_path: str):
    """检查 Excel 文件是否存在且扩展名合法。"""
    return ensure_excel_file_common(file_path)


def _get_active_worksheet(workbook: Workbook) -> Worksheet:
    active = workbook.active
    if active is None:
        raise ToolValidationError("Workbook has no active sheet")
    return cast(Worksheet, active)


def load_workbook_safe(file_path: str, data_only: bool = True) -> Workbook:
    """安全加载工作簿。

    参数:
    - data_only=True: 读取公式计算后的值，而不是公式字符串
    """
    ensure_excel_file(file_path)

    try:
        workbook = load_workbook(file_path, data_only=data_only)
    except Exception as exc:  # noqa: BLE001
        raise ToolValidationError(f"Failed to load Excel workbook: {exc}") from exc

    return workbook


def get_sheet_names(file_path: str) -> List[str]:
    """返回工作簿中所有 sheet 名。"""
    workbook = load_workbook_safe(file_path)
    return list(workbook.sheetnames)


def get_active_sheet_name(file_path: str) -> str:
    """返回当前 active sheet 的名称。"""
    workbook = load_workbook_safe(file_path)
    return _get_active_worksheet(workbook).title


def get_worksheet(
    file_path: str,
    sheet_name: Optional[str] = None,
    sheet_index: Optional[int] = None,
) -> Worksheet:
    """按 sheet_name 或 sheet_index 获取工作表。

    规则：
    - 优先使用 sheet_name
    - 其次使用 sheet_index
    - 都不传则返回 active sheet
    """
    workbook = load_workbook_safe(file_path)

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ToolValidationError(f"Sheet not found: {sheet_name}")
        return cast(Worksheet, workbook[sheet_name])

    if sheet_index is not None:
        if sheet_index < 0 or sheet_index >= len(workbook.worksheets):
            raise ToolValidationError(
                f"sheet_index out of range: {sheet_index}, "
                f"available: 0 ~ {len(workbook.worksheets) - 1}"
            )
        return cast(Worksheet, workbook.worksheets[sheet_index])

    return _get_active_worksheet(workbook)


def normalize_cell_value(value: Any, trim_strings: bool = True) -> Any:
    """标准化单元格值。"""
    if isinstance(value, str) and trim_strings:
        return value.strip()
    return value


def read_row_values(
    file_path: str,
    row_index: int,
    sheet_name: Optional[str] = None,
    sheet_index: Optional[int] = None,
    trim_strings: bool = True,
) -> List[Any]:
    """读取某一行的所有值。"""
    if row_index < 1:
        raise ToolValidationError("row_index must be >= 1")

    sheet = get_worksheet(file_path, sheet_name=sheet_name, sheet_index=sheet_index)

    for row in sheet.iter_rows(min_row=row_index, max_row=row_index, values_only=True):
        return [normalize_cell_value(cell, trim_strings=trim_strings) for cell in row]

    return []


def build_headers(raw_headers: List[Any]) -> List[str]:
    """将原始表头标准化为字符串列表，并处理空列/重名列。"""
    headers: List[str] = []
    seen: Dict[str, int] = {}

    for idx, value in enumerate(raw_headers, start=1):
        text = "" if value is None else str(value).strip()

        if not text:
            text = f"__EMPTY_COL_{idx}"

        if text in seen:
            seen[text] += 1
            text = f"{text}__{seen[text]}"
        else:
            seen[text] = 1

        headers.append(text)

    return headers


def get_headers(
    file_path: str,
    header_row: int = 1,
    sheet_name: Optional[str] = None,
    sheet_index: Optional[int] = None,
    trim_strings: bool = True,
) -> List[str]:
    """读取表头行并标准化。"""
    raw_headers = read_row_values(
        file_path=file_path,
        row_index=header_row,
        sheet_name=sheet_name,
        sheet_index=sheet_index,
        trim_strings=trim_strings,
    )
    return build_headers(raw_headers)


def is_empty_row(values: List[Any]) -> bool:
    """判断一整行是否为空。"""
    for value in values:
        if value is None:
            continue
        if isinstance(value, str) and value.strip() == "":
            continue
        return False
    return True


def read_sheet_as_records(
    file_path: str,
    sheet_name: Optional[str] = None,
    sheet_index: Optional[int] = None,
    header_row: int = 1,
    start_row: Optional[int] = None,
    max_rows: Optional[int] = None,
    drop_empty_rows: bool = True,
    trim_strings: bool = True,
) -> List[Dict[str, Any]]:
    """把工作表读取为 records 结构。

    输出示例:
    [
        {"姓名": "张三", "班级": "一班"},
        {"姓名": "李四", "班级": "二班"},
    ]
    """
    if header_row < 1:
        raise ToolValidationError("header_row must be >= 1")

    if start_row is not None and start_row < 1:
        raise ToolValidationError("start_row must be >= 1")

    if max_rows is not None and max_rows < 1:
        raise ToolValidationError("max_rows must be >= 1")

    sheet = get_worksheet(file_path, sheet_name=sheet_name, sheet_index=sheet_index)
    headers = get_headers(
        file_path=file_path,
        header_row=header_row,
        sheet_name=sheet_name,
        sheet_index=sheet_index,
        trim_strings=trim_strings,
    )

    data_start_row = start_row if start_row is not None else header_row + 1
    records: List[Dict[str, Any]] = []
    count = 0

    for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
        values = [normalize_cell_value(v, trim_strings=trim_strings) for v in row]

        if len(values) < len(headers):
            values.extend([None] * (len(headers) - len(values)))
        elif len(values) > len(headers):
            values = values[:len(headers)]

        if drop_empty_rows and is_empty_row(values):
            continue

        record = dict(zip(headers, values))
        records.append(record)

        count += 1
        if max_rows is not None and count >= max_rows:
            break

    return records


def inspect_workbook(file_path: str) -> Dict[str, Any]:
    """返回工作簿级别的结构信息。"""
    workbook = load_workbook_safe(file_path)
    active_sheet = _get_active_worksheet(workbook)

    return {
        "file_path": file_path,
        "sheet_names": list(workbook.sheetnames),
        "sheet_count": len(workbook.sheetnames),
        "active_sheet": active_sheet.title,
    }


def inspect_worksheet(
    file_path: str,
    sheet_name: Optional[str] = None,
    sheet_index: Optional[int] = None,
    header_row: int = 1,
    sample_rows: int = 5,
) -> Dict[str, Any]:
    """返回单个工作表的结构信息。"""
    if sample_rows < 1:
        raise ToolValidationError("sample_rows must be >= 1")

    sheet = get_worksheet(file_path, sheet_name=sheet_name, sheet_index=sheet_index)
    headers = get_headers(
        file_path=file_path,
        header_row=header_row,
        sheet_name=sheet.title,
    )
    sample_records = read_sheet_as_records(
        file_path=file_path,
        sheet_name=sheet.title,
        header_row=header_row,
        start_row=header_row + 1,
        max_rows=sample_rows,
        drop_empty_rows=True,
        trim_strings=True,
    )

    return {
        "file_path": file_path,
        "sheet_name": sheet.title,
        "max_row": sheet.max_row,
        "max_column": sheet.max_column,
        "header_row": header_row,
        "headers": headers,
        "sample_records": sample_records,
    }